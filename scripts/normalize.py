import pandas as pd
import sys
import os
import re

from openpyxl import load_workbook

# ======================================================
# DETECT REPORT
# ======================================================
def detect_report(file_name):
    normalized = file_name.lower()
    if '000019' in normalized or 'lk-000019' in normalized:
        return 'LK-000019'
    elif '000035' in normalized or 'lk-000035' in normalized:
        return 'LK-000035'
    elif '000042' in normalized or 'lk-000042' in normalized:
        return 'LK-000042'
    elif '000117' in normalized or 'lk-000117' in normalized:
        return 'LK-000117'
    elif '000040' in normalized or 'lk-000040' in normalized:
        return 'LK-000040'
    elif '000064' in normalized or 'lk-000064' in normalized:
        return 'LK-000064'
    elif '000104' in normalized or 'lk-000104' in normalized:
        return 'LK-000104'
    elif '000105' in normalized or 'lk-000105' in normalized:
        return 'LK-000105'
    elif '000106' in normalized or 'lk-000106' in normalized:
        return 'LK-000106'
    elif '000107' in normalized or 'lk-000107' in normalized:
        return 'LK-000107'
    elif '000137' in normalized or 'lk-000137' in normalized:
        return 'LK-000137'
    elif '000143' in normalized or 'lk-000143' in normalized:
        return 'LK-000143'
    elif '000148' in normalized or 'lk-000148' in normalized:
        return 'LK-000148'
    elif '000149' in normalized or 'lk-000149' in normalized:
        return 'LK-000149'
    elif '000150' in normalized or 'lk-000150' in normalized:
        return 'LK-000150'
    elif '000155' in normalized or 'lk-000155' in normalized:
        return 'LK-000155'
    elif '000167' in normalized or 'lk-000167' in normalized:
        return 'LK-000167'
    elif '000020' in normalized or 'lk-000020' in normalized:
        return 'LK-000020'
    elif '000021' in normalized or 'lk-000021' in normalized:
        return 'LK-000021'
    elif '000030' in normalized or 'lk-000030' in normalized:
        return 'LK-000030'
    elif '000031' in normalized or 'lk-000031' in normalized:
        return 'LK-000031'
    elif '000045' in normalized or 'lk-000045' in normalized:
        return 'LK-000045'
    elif '000075' in normalized or 'lk-000075' in normalized:
        return 'LK-000075'
    elif '000109' in normalized or 'lk-000109' in normalized:
        return 'LK-000109'
    elif '000115' in normalized or 'lk-000115' in normalized:
        return 'LK-000115'
    elif '000124' in normalized or 'lk-000124' in normalized:
        return 'LK-000124'
    elif '000151' in normalized or 'lk-000151' in normalized:
        return 'LK-000151'
    elif '000075' in normalized or 'lk-000075' in normalized:
        return 'LK-000075'
    elif '000014' in normalized or 'lk-000014' in normalized:
        return 'LK-000014'
    elif '000048' in normalized or 'lk-000048' in normalized:
        return 'LK-000048'
    elif '000065' in normalized or 'lk-000065' in normalized:
        return 'LK-000065'
    elif '000153' in normalized or 'lk-000153' in normalized:
        return 'LK-000153'
    elif '000010' in normalized or 'lk-000010' in normalized:
        return 'LK-000010'
    return 'UNKNOWN'

# ======================================================
# EXPECTED HEADER
# ======================================================
EXPECTED_COLUMNS = {
    'LK-000019': [
        'intp', 'inno', 'insl', 'slna', 'intc', 'inco', 'inna', 'chan', 'tipe', 
        'npwp', 'inal', 'intl', 'inct', 'noso', 'tcso', 'pl', 'co', 'na', 'ns', 
        'un', 'uk', 'u1', 'u2', 'u3', 'qt', 'q1', 'q2', 'q3', 'hd', 'hl', 'hs', 
        'dc1', 'dr1', 'dc2', 'dr2', 'dc3', 'dr3', 'dc4', 'dr4', 'dcrp', 'dcrpt', 
        'tt', 'indr', 'inpr', 'innm', 'copc', 'inlo', 'pgrup', 'pgnama', 'inc2', 'inc3',
        'kodetransaksi', 'tanggal', 'tglkirim', 'jatuhtempo', 'referensi', 'kodecabang',
        'kodedivisi', 'namadivisi', 'kodeorder', 'kodecustomer', 'namacustomer', 'alamat',
        'kodesalesman', 'namasalesman', 'tipeharga', 'kodegudang', 'term', 'keteranganfaktur',
        'totalbruto', 'diskonglobalclaimpersen', 'diskonglobalclaimrupiah', 'diskonglobalregpersen',
        'diskonglobalregrupiah', 'tarifppn', 'totalpajak', 'totalnetto', 'totaldpp', 'totalppn',
        'jenisppn', 'nopajak', 'tglfakturpajak', 'cetakke', 'noninsentif', 'fakturkembali',
        'nourut', 'kodebarang', 'kodepalet', 'namabarang', 'supplier', 'qty4', 'satuan4',
        'qty3', 'satuan3', 'qty2', 'satuan2', 'qty1', 'satuan1', 'totqty', 'totalsatuan',
        'hargasatuan', 'hargalist', 'bruto', 'diskonitemclaimpersen', 'diskonitemclaimrupiah',
        'diskonitemregpersen', 'diskonitemregrupiah', 'diskonbarang', 'jumlah', 'keteranganitem',
        'discglobalitemclaim', 'discglobalitemreg', 'netto', 'dpp', 'ppn', 'total',
        'kd_supplier', 'kd_barang', 'kd_paret', 'nama_barang', 'kd_cabang', 'kd_gudang',
        'so_akhir', 'satuan', 'harga', 'jumlah', 'isi4', 'isi3', 'isi2',
    ],
    'LK-000035': [
        'nama pemasok utama barang',
        'nama penjual',
        'tgl faktur',
        'no. faktur',

        'no. pelanggan',
        'nama pelanggan',
        'total',

        'alamat 1 pelanggan',
        'kota pelanggan',
        'nama tipe pelanggan pelanggan',
        'tipe outlet',
        'no. barang',
        'kode yuri',
        'keterangan barang',
        'kuantitas',
        'jumlah',
        'harga satuan barang',
        'inc ppn 11%',
        'in ctn',
        'bulan',
        'produk k1',
        'produk k2',

        'deskripsi barang',
        'status',
        'bulan 1',
        'bulan 2',
        'bulan 3',
        'bulan 4',
        'bulan 5',
        'bulan 6',
        'avg 3m',
        'avg 6m',
        'bulan berjalan',
        'berjalan/avg',
        'stok on hand',
        'r. order',
        'doi',
        'harga satuan',
        'value stok',
        'value avg 6m'
    ],
    'LK-000042': [
        'kode customer', 'nama customer', 'alamat customer', 'telp customer', 'sales',
        'invoice', 'tgl invoice', 'type customer', 'daerah', 'kode', 'kode prinsipal',
        'nama barang', 'qty', 'satuan', 'bonus', 'harga', 'disc', 'tax', 'd1', 'd2',
        'd3', 'total netto', 'kode principle', 'ukuran', 'sat', 'isi', 'qty in ct'
    ],
    'LK-000117': [
        'divisi', 'product grup level 3', 'product code', 'product name', 'stock',
        'distributor', 'cabang', 'tipetrans', 'principal', 'productgroup1', 'productgroup2',
        'productgroup3', 'brand', 'kodeproduk', 'kodevarian', 'kodeprodukprincipal',
        'namaproduk', 'packaging', 'productclass', 'kodecustomer', 'namacustomer',
        'alamatcustomer', 'area', 'subarea', 'channel', 'subchannel', 'customergroup',
        'keyaccount', 'kodesalesman', 'namasalesman', 'kodesalesco', 'namasalesco',
        'kodespv', 'namaspv', 'tahun-bulan', 'bulan', 'tanggal', 'weekno', 'nomornota',
        'salemethod', 'sellingtype', 'qtysold', 'qtysoldcrt', 'qtysolduom1', 'qtysolduom2',
        'qtysolduom3', 'qtysolduom4', 'qtysoldtotalpcs', 'freegoodtotalpcs', 'tonnage',
        'volume(ltr)', 'grossamount', 'linediscount1', 'linediscount2', 'linediscount3',
        'linediscount4', 'linediscount5', 'totallinediscount', 'discountnota1',
        'discountnota2', 'discountnota3', 'totaldiscountnota', 'dpp', 'ppn', 'ppnbm',
        'tax3', 'netamount', 'warehouse', 'customerpo', 'customerjoindate', 'nofakturpajak',
        'tglfakturpajak', 'nomorfakturproforma', 'tglfakturproforma', 'term', 'uom1',
        'uom2', 'uom3', 'uom4', 'isiuom1', 'isiuom2', 'isiuom3', 'sellingprice', 'cogs',
        'sellingpriceinkg', 'caseweightinkg', 'qtyordertotalpcs', 'tslqtysoldnfg',
        'tslconvpcstoctn', 'tsltonnagesoldfg'
    ],
    'LK-000040': [
        'no',
        'kode',
        'nama barang',
        'sisa (ac)',
        'penjua_id',
        'tanggal',
        'no',
        'pelang',
        'kode',
        'nama',
        'alamat',
        'pj_har_lst',
        'sales',
        'sales_kom',
        'stock_id',
        'group_st2',
        'kode_c',
        'nama',
        'namaex1',
        'jumlah',
        'bon_jumlah',
        'harga',
        'disc_pers',
        'disc_persn',
        'd2',
        'mast_disc',
        'nilai'
    ],
    
    'LK-000064': [
        'no surat jalan',
        'no system',
        'kode toko',
        'kode sales',
        'tanggal',
        'no urut',
        'kode produk',
        'qty jual terkecil',
        'harga terkecil',
        'disc 1',
        'disc 2',
        'disc 3',
        'total bruto',
        'disc 1 rp',
        'disc 2 rp',
        'disc 3 rp',
        'total netto',
        'kode supplier',
        'nama barang',
        'isi barang',
        'nama toko',
        'alamat',
        'tipe toko',
        'npwp',
        'no ktp',
        'kode',
        'nama salesman',
        'kode barang',
        'nama barang',
        'kode barcode',
        'stok akhir',
        'isi besar'
    ],
    
    'LK-000104': [
        'no. faktur',
        'tgl faktur',
        'no. pelanggan',
        'nama pelanggan',
        'alamat pelanggan',
        'no. barang',
        'keterangan barang',
        'qty',
        'unit 1 brg',
        'harga satuan',
        'dpp',
        'ppn',
        'jumlah',
        'type pelanggan',
        'nama penjual',
        'nama kategori barang',
        'no. barang',
        'deskripsi barang',
        'rasio 2',
        'harga satuan',
        'kts. stok',
        'unit 1',
        'kts dalam unit 2'
    ],
    
    'LK-000105': [
        'no',
        'order date',
        'jenis outlet',
        'cd outlet',
        'nama outlet',
        'so num',
        'kode produk',
        'nama produk',
        'krt',
        'pack',
        'pcs',
        'bonus',
        'gross. sales',
        'disc 1',
        'disc 2',
        'disc 3',
        'disc 4',
        'disc 5',
        'bonus',
        'total diskon',
        'ppn',
        'gross',
        'retur',
        'netto',
        'id sales',
        'sales',
        'cabang',
        'no',
        'cd product',
        'nm product',
        'warehouse',
        'avl qty (krt)',
        'avl qty (pack)',
        'avl qty (pcs)',
        'tot available (pcs)',
        'avl nilai(rp)',
    ],
    
    'LK-000106': [
        'tanggal',
        'no. dokumen',
        'surat jalan',
        'outlet code',
        'nama customer',
        'salesman',
        'alamat',
        'code',
        'nama barang',
        'penjualan karton',
        'penjualan lsn',
        'penjualan pcs',
        'discount (%)',
        'grand total ppn',
        'item no.',
        'item description',
        'unit 1',
        'centre cv. ap',
        'm-centre cv. apm',
    ],
    'LK-000107': [
        'nama distributor',
        'kode pelanggan',
        'nama pelanggan',
        'alamat',
        'nomor fj',
        'tgl fj',
        'kode barang',
        'nama barang',
        'nama salesman',
        'qty kecil',
        'gross amt',
        'est total net',
        'tipe transaksi',
        'kd item',
        'kd item principle',
        'item',
        'brand',
        'gudang',
        'base qty',
        'price jual',
        'price beli',
        'qty ctn',
        'qty pcs',
        'value',
        'value + ppn',
    ],
    
    'LK-000137': [
        'distributor',
        'trans',
        'kd_produk',
        'kd_pabrik',
        'nm_produk',
        'subbrand(group)',
        'brand',
        'dept',
        'principal',
        'kategori',
        'prod_class',
        'qty',
        'kartonutuh',
        'pieces',
        'freegood_pcs',
        'nilai_brutto_dpp',
        'nilai_netto_dpp',
        'nilai_netto_ppn',
        'nilai_freegood',
        'discitem',
        'discnota',
        'kd_cust',
        'nm_cust',
        'type',
        'area',
        'subarea',
        'salesman',
        'supervisor',
        'hcode1',
        'hdesc1',
        'hcode2',
        'hdesc2',
        'hcode3',
        'hdesc3',
        'no_nota',
        'tanggal',
        'bulan',
        'tahun-bulan',
        'week',
        'keyaccount',
        'kodesalesman',
        'custgroup',
        'cabang',
        'customerpo',
        'divisi',
        'tonnage',
        'qtysold_pcs',
        'discspecial',
        'discspecialrp',
        'discpromositotalrp',
        'joindatesalesman',
        'joindatecustomer',
        'product_barcode',
        'customer_barcode',
        'driverid',
        's126',
        'pt. joenoes ikamulya',
    ],
    
    'LK-000143': [
        'customer name',
        'customer#',
        'address',
        'area',
        'product code',
        'product name',
        'packaging',
        'varian',
        'invoice date',
        'invoice no',
        'salesorder#',
        'category',
        'salesman',
        'uom',
        'quantity',
        'qty (pcs)',
        'qty (ctn)',
        'freegood',
        'price',
        'gross amount',
        'linedisc1',
        'linedisc2',
        'linedisc3',
        'linedisc4',
        'linedisc5',
        'ld amount',
        '%disc1',
        '%disc2',
        '%disc3',
        'discount',
        'dpp',
        'tax',
        'net amount',
        'divisi',
        'product grup level 3',
        'product code',
        'product name',
        'packaging',
        'uom',
        'stock',
        'stock (pcs)',
        'stock (ctn)',
        'tonnage',
        'volume',
        'stock uom1',
        'stock uom2',
        'stock uom3',
        'value@selling',
    ],
    
    'LK-000148': [
        'no. urut',
        'kode brg',
        'nama brg',
        'sat std',
        'sat trans',
        'kode jns brg',
        'nama jns brg',
        'kode grp brg',
        'nama grp brg',
        'saldo awal',
        'pb',
        'lpb',
        'nrb',
        'total pembelian',
        'apotek',
        'toko obat',
        'toko kelontong',
        'toko kosmetik',
        'grosir',
        'market',
        'lain2',
        'nrj',
        'total penjualan',
        'bs',
        'koreksi stok',
        'tbl',
        'kbl',
        'mtb',
        'saldo akhir',
        'no. urut',
        'ref',
        'tgl ref',
        'no. ref',
        'kode cust',
        'nama cust',
        'kode sales',
        'nama sales',
        'keterangan',
        'kode',
        'nama',
        'nama jns brg',
        'nama div pabrik',
        'nama grp brg',
        'qty',
        'sat',
        'harga unit(mu)',
        'total(mu)',
        'subtot net pjk(mu)',
        'nama kota',
    ],
    
    'LK-000149': [
        'no. urut',
        'kode brg',
        'nama brg',
        'sat std',
        'sat trans',
        'kode jns brg',
        'nama jns brg',
        'kode grp brg',
        'nama grp brg',
        'saldo awal',
        'pb',
        'lpb',
        'nrb',
        'total pembelian',
        'apotek',
        'toko obat',
        'toko kelontong',
        'toko kosmetik',
        'grosir',
        'market',
        'lain2',
        'nrj',
        'total penjualan',
        'bs',
        'koreksi stok',
        'tbl',
        'kbl',
        'mtb',
        'saldo akhir',
        'no. urut',
        'ref',
        'tgl ref',
        'no. ref',
        'kode cust',
        'nama cust',
        'kode sales',
        'nama sales',
        'keterangan',
        'kode',
        'nama',
        'nama jns brg',
        'nama div pabrik',
        'nama grp brg',
        'qty',
        'sat',
        'harga unit(mu)',
        'total(mu)',
        'subtot net pjk(mu)',
        'nama kota',
    ],
    
    'LK-000150': [
        'no. urut',
        'ref',
        'tgl ref',
        'no. ref',
        'kode cust',
        'nama cust',
        'kode sales',
        'nama sales',
        'keterangan',
        'kode',
        'nama',
        'nama jns brg',
        'nama div pabrik',
        'nama grp brg',
        'qty',
        'sat',
        'harga unit(mu)',
        'total(mu)',
        'subtot net pjk(mu)',
        'nama kota',
        'qty in ctn',
        'bulan',
        'produk k1',
        'produk k2',
        'no. urut',
        'kode brg',
        'nama brg',
        'sat std',
        'sat trans',
        'kode jns brg',
        'nama jns brg',
        'kode grp brg',
        'nama grp brg',
        'saldo awal',
        'saldo akhir',
        'in ctn',
    ],
    
    'LK-000155': [
        'kode toko',
        'nama toko',
        'alamat toko',
        'kecamatan',
        'tipe toko',
        'kode sales',
        'nama sales',
        'supervisor',
        'tipe transaksi',
        'tgl faktur',
        'no faktur',
        'grup barang',
        'kode barang',
        'kode lama barang',
        'nama barang',
        'nama supplier',
        'kode unit',
        'qty',
        'harga (rp)',
        'jml bruto',
        'inv. discount',
        'discount',
        'net',
        'ppn',
        'cogs',
        'total',
        'disc.1',
        'disc.2',
        'disc.3',
        'disc.4',
        'disc.5',
        'disc.1',
        'disc.2',
        'disc.3',
        'disc.4',
        'disc.5',
        'pot. rupiah',
        'entity',
        'site',
        'channel',
        'manual ref',
        'nama supplier',
        'item id',
        'keterangan item',
        'qty ctn',
        'unit',
        'qty pcs',
        'unit',
        'ccy',
        'harga jual',
        'harga beli',
        'nilai',
        'supplier 2',
    ],
    
    'LK-000167': [
        'pemasok',
        'kode barang',
        'upc/barcode',
        'nama barang',
        'nama gudang',
        'isi',
        'stock akhir',
        'no invoice',
        'tanggal',
        'jatuh tempo',
        'bulan',
        'tahun',
        'id pelanggan',
        'pelanggan',
        'npwp pelanggan',
        'alamat faktur',
        'kelurahan',
        'kecamatan',
        'kota alamat',
        'jenis market',
        'group market',
        'id karyawan penjual',
        'nama penjual utama faktur',
        'kode #',
        'nama barang',
        'kuantitas',
        'nama sa',
        '@harga',
        'diskon',
        'diskon %',
        'ppn barang',
        'total harga',
        'nama gudang',
        'nama pemasok utama barang & jasa',
        'keterangan faktur penjualan',
        'tanggal penerimaan',
    ],
    'LK-000020': [
        'no',
        'divisi',
        'tanggal',
        'nomor nota',
        'j.tempo',
        'nama salesman',
        'area pelanggan',
        'tipe pelanggan',
        'kode pelanggan',
        'nama pelanggan',
        'alamat',
        'kode barang',
        'nama barang',
        'quantity sat-1',
        'quantity sat-2',
        'quantity sat-3',
        'bonus sat-1',
        'bonus sat-2',
        'bonus sat-3',
        'qty satuan lengkap',
        'bns satuan lengkap',
        'harga / sat',
        'jumlah-rp',
        'discount (1)-rp',
        'discount (2)-rp',
        'discount (3)-rp',
        'discount (4)-rp',
        'discount (5)-rp',
        'netto',
        'ppn',
        'netto + ppn',
        'no',
        'group',
        'kode',
        'nama barang',
        'stock',
        'stock akhir',
        'harga',
        'jumlah-rp',
    ],
    
    'LK-000021': [
        'nama_agen',
        'kode_customer',
        'nama_customer',
        'alamat_customer',
        'invoice_nomor_agen',
        'tanggal_invoice',
        'tipe_customer',
        'kota',
        'sku_kode_agen',
        'nama_sku',
        'pcs',
        'diskon_1_persen',
        'diskon_2_persen',
        'diskon_3_persen',
        'diskon_4_persen',
        'diskon_5_persen',
        'diskon_6_persen',
        'quantity_bonus',
        'satuan',
        'rafraksi_rp',
        'total_invoice_value',

        'no',
        'kode_sku_agen',
        'kode_sku_jim',
        'nama_sku_jim',
        'isi_per_carton',
        'qty_carton',
    ],
    
    'LK-000030': [
        'periode',
        'no_faktur',
        'tanggal',
        'jatuh_tempo',
        'kode_customer',
        'nama_customer',
        'alamat',
        'tipe_customer',
        'nama_tipe',
        'kode_sales',
        'nama_sales',
        'nu',
        'kode_barang',
        'nama_barang',
        'kode_grup',
        'nama_grup',
        'qty',
        'satuan',
        'harga',
        'jumlah',
        'disc_bp1',
        'disc_bp2',
        'disc_bp3',
        'disc_bu3',
        'disc_barang',
        'disc_faktur',
        'ppn',
        'netto',
        'cash',
        'credit',
        'keterangan',

        'kode_supplier',
        'nama_supplier',
        'kode_grup',
        'nama_grup',
        'kode_barang',
        'nama_barang',
        'satuan_1',
        'satuan_2',
        'satuan_3',
        'ratio_1',
        'ratio_2',
        'ratio_3',
        'qty_1',
        'qty_2',
        'qty_3',
        'jumlah',
    ],
    
    'LK-000031': [
        'tgl_faktur',
        'no_faktur',
        'kode_customer',
        'nama_customer',
        'kode_pos',
        'area_customer',
        'tipe_customer',
        'kode_sales',
        'nama_sales',
        'kode_barang',
        'kode_barang_2',
        'nama_barang',
        'satuan',
        'harga',
        'qty_barang',
        'diskon',
        'net_value',
        'alasan',
        'id_pabrik',
        'nama_group',
        'ppn',
        'net_value_with_ppn',
        'id_stock',
        'kode_barang',
        'nama_barang',
        'id_division',
        'pabrik',
        'id_brand',
        'qty',
        'qty_in_cs',
        'block',
        'qtyblock_in_cs',
        'intransit',
        'intransit_in_cs',
        'bad',
        'bad_in_cs',
        'qty_temp',
        'qtytemp_in_cs',
        'good_skw',
        'good_skw_in_cs',
        'good_ktp',
        'good_ktp_in_cs',
        'good_stg',
        'good_stg_in_cs',
        'kanvas',
        'kanvas_in_cs',
        'satuan',
        'harga',
        'total',
        'uom_cs',
    ],
    
    'LK-000045': [
        'tanggal',
        'no_faktur',
        'kode_salesman',
        'nama_salesman',
        'kode_customer',
        'nama_customer',
        'alamat_customer',
        'kelompok_barang',
        'kode_barang',
        'nama_barang',
        'qty_pcs',
        'isi_tengah',
        'harga_jual',
        'diskon_reguler',
        'pds_1',
        'pds_2',
        'pds_3',
        'brutto',
        'netto',

        'kode_barang',
        'nama_gudang',
        'nama_barang',
        'kelompok_barang',
        'ctn',
        'total_pcs',
        'isi_tengah',
        'isi_besar',
        'isi_besar_2',
    ],
    
    'LK-000075': [
        'transaksi',
        'kode_produk',
        'nama_produk',
        'grup_produk',
        'principal',
        'product_class',
        'qty',
        'retur_qty',
        'karton_utuh',
        'pieces',
        'freegood_pcs',
        'nilai_bruto_dpp',
        'nilai_netto_dpp',
        'nilai_bruto_ppn',
        'nilai_netto_ppn',
        'harga_per_pcs',
        'nilai_freegood',
        'reg',
        'reg_1',
        'reg_rp_1',
        'reg_2',
        'reg_rp_2',
        'reg_3',
        'reg_rp_3',
        'reg_4',
        'reg_rp_4',
        'reg_5',
        'reg_rp_5',
        'total_reg_rp',
        'ext',
        'ext_1',
        'ext_rp_1',
        'ext_2',
        'ext_rp_2',
        'ext_3',
        'ext_rp_3',
        'ext_4',
        'ext_rp_4',
        'ext_5',
        'ext_rp_5',
        'total_ext_rp',
        'kode_customer',
        'nama_customer',
        'tipe_customer',
        'area',
        'sub_area',
        'salesman',
        'no_nota',
        'tanggal',
        'bulan',
        'tahun',
        'minggu',
        'kode_salesman',
        'alamat',

        'no',
        'kode_produk',
        'alias',
        'keterangan',
        'isi_besar',
        'satuan_besar',
        'harga_beli_besar',
        'isi_kecil',
        'satuan_kecil',
        'harga_beli_kecil',
        'total_qty',
        'nilai',
    ],
    
    'LK-000109': [
        'no',
        'group',
        'kode_barang',
        'nama_barang',
        'stock',
        'stock_akhir',
        'harga',
        'jumlah_rp',
        'no',
        'divisi',
        'tanggal',
        'nomor_nota',
        'jatuh_tempo',
        'nama_salesman',
        'area_pelanggan',
        'tipe_pelanggan',
        'kode_pelanggan',
        'nama_pelanggan',
        'alamat',
        'kode_barang',
        'nama_barang',
        'qty_sat_1',
        'qty_sat_2',
        'qty_sat_3',
        'bonus_sat_1',
        'bonus_sat_2',
        'bonus_sat_3',
        'qty_satuan_lengkap',
        'bonus_satuan_lengkap',
        'harga_satuan',
        'jumlah_rp',
        'discount_1_rp',
        'discount_2_rp',
        'discount_3_rp',
        'discount_4_rp',
        'discount_5_rp',
        'netto',
        'ppn',
        'netto_ppn',
    ],
    
    'LK-000115': [
        'kode_barang',
        'nama_barang',
        'kode_supplier',
        'nama_supplier',
        'qty',
        'isi_besar',
        'isi_tengah',
        'jumlah',
        'jumlah_brutto',
        'jumlah_discount',
        'jumlah_netto_ex',
        'jumlah_ppn',
        'kode_customer',
        'nama_customer',
        'kota_customer',
        'hari_kunjungan',
        'pola_kunjungan',
        'tipe_customer',
        'kode_salesman',
        'nama_salesman',
        'area',
        'tanggal',
        'no_faktur',
        'term',
        'alamat_customer',
        'kelompok_barang',
        'pdiscount_item',
        'pdiscount_item_2',
        'no_batch',
        'tgl_exp',
        'bulan',
        'nama_gudang',
        'rp_discount_item',
        'rp_discount_item_2',
        'pdiscount_item_3',
        'rp_discount_item_3',
        'kode_sarana',
        'kode_obat',
        'pdiscount_item_4',
        'rp_discount_item_4',
        'subtipe_customer',
        'kode_principle',
        'kode_lokasi',
        'jenis',
        'no_sp',
        'cabang_principle',
        'catatan_invoice',
        'user_id',
        'tgl_sp',
        'subtipe_sup',
        'nama_barang',
        'sisa',
        'kelompok',
        'nama_gudang',
        'kode_barang',
        'satuan_kecil',
        'kelompok_2',
        'kode_obat',
        'satuan_kemasan',
    ],
    
    'LK-000124': [
        'no',
        'faktur_code',
        'faktur_date',
        'due_date',
        'supplier_code',
        'supplier_name',
        'kategori_name',
        'barang_code',
        'barang_name',
        'customer_code',
        'customer_name',
        'klasifikasi_name',
        'customer_address',
        'customer_kota',
        'wilayah_name',
        'salesperson_name',
        'qty_besar',
        'satuan_besar',
        'harga_satuan_besar',
        'qty_kecil',
        'satuan_kecil',
        'harga_satuan_kecil',
        'qty_bonus',
        'qty_potongan_stok',
        'subtotal',
        'disc_persen_1',
        'disc_persen_2',
        'disc_persen_3',
        'disc_persen_4',
        'total_discount',
        'total_sebelum_pajak',
        'ppn_rp',
        'total',
        'item_code_pt_jim',
        'item_name',
        'item_per_box',
        'item_code_agen',
        'stock_in_karton',
    ],
    
    'LK-000151': [
        'no',
        'barang_code',
        'barang_name',
        'satuan_standar',
        'satuan_transaksi',
        'jenis_barang_code',
        'jenis_barang_name',
        'group_barang_code',
        'group_barang_name',
        'saldo_awal',
        'pb',
        'lpb',
        'nrb',
        'total_pembelian',
        'apotek',
        'toko_obat',
        'toko_kelontong',
        'toko_kosmetik',
        'grosir',
        'market',
        'lain_lain',
        'nrj',
        'total_penjualan',
        'bs',
        'koreksi_stok',
        'tbl',
        'kbl',
        'mtb',
        'saldo_akhir',
        'ref',
        'ref_date',
        'ref_no',
        'customer_code',
        'customer_name',
        'salesperson_code',
        'salesperson_name',
        'description',
        'barang_code',
        'barang_name',
        'jenis_barang_name',
        'divisi_pabrik_name',
        'group_barang_name',
        'qty',
        'satuan',
        'harga_unit',
        'total',
        'subtotal_net_pajak',
        'customer_kota',
        'qty_in_ctn',
        'bulan',
    ],
    
    'LK-000075' : [
        'no',
        'barang_code',
        'barang_alias',
        'barang_name',
        'isi',

        'qty_besar',
        'satuan_besar',
        'harga_beli_besar',

        'qty_kecil',
        'satuan_kecil',
        'harga_beli_kecil',

        'total_qty',
        'nilai',

        'trans_type',
        'barang_code',
        'barang_name',
        'group_name',
        'principal_name',
        'product_class',

        'qty',
        'retur_qty',
        'karton_utuh',
        'pieces',
        'freegood_pcs',

        'nilai_bruto_dpp',
        'nilai_netto_dpp',
        'nilai_bruto_ppn',
        'nilai_netto_ppn',

        'harga_per_pcs',
        'nilai_freegood',

        'reg',
        'reg1',
        'reg_rp1',
        'reg2',
        'reg_rp2',
        'reg3',
        'reg_rp3',
        'reg4',
        'reg_rp4',
        'reg5',
        'reg_rp5',
        'total_reg',

        'ext',
        'ext1',
        'ext_rp1',
        'ext2',
        'ext_rp2',
        'ext3',
        'ext_rp3',
        'ext4',
        'ext_rp4',
        'ext5',
        'ext_rp5',
        'total_ext',

        'customer_code',
        'customer_name',
        'customer_type',
        'area_name',
        'subarea_name',

        'salesperson_name',
        'salesperson_code',

        'invoice_no',
        'invoice_date',

        'bulan',
        'tahun',
        'minggu',

        'customer_address',
    ],
    
    'LK-000014' : [
        'no_barang',
        'deskripsi_barang',
        'baik',
        'nama_kategori_barang',

        'no_faktur',
        'tgl_faktur',

        'no_pelanggan',
        'nama_tipe_pelanggan',

        'nama_pelanggan',
        'alamat_1_pelanggan',

        'nama_penjual',

        'no_barang',
        'keterangan_detil',

        'kts_faktur',
        'satuan',
        'kuantitas',
        'unit_1_barang',

        'harga_satuan',
        'jumlah',

        'persen_diskon',
        'diskon_barang',

        'dpp_ppn',
        'code_tax1',

        'nama_dept',
    ],
    
    'LK-000048' : [
        'kode_kategori',
        'nama_kategori',
        'kode_barang',
        'nama_barang',
        'qty',
        'satuan',
        'qty2',
        'satuan2',
        'jumlah',
        'harga',
        'nilai',

        'kode_sales',
        'nama_sales',
        'kode_customer',
        'nama_customer',
        'nama_kategori',
        'satuan_netto',
        'kode_barang',
        'nama_barang',
        'tgl',

        'qty_jual',
        'qty_extra',

        'harga_jual',
        'nilai_jual',
        'ppn_jual',
        'netto_jual',

        'qty_retur',
        'harga_retur',
        'nilai_retur',
        'ppn_retur',
        'netto_retur',

        'sales_luar',
        'no_trans',

        'nama_area',
        'nama_segment',

        'qty_pcs',
        'harga_pcs',

        'no_so',
        'id_pelanggan',
        'id_lot',
        'exp_date',

        'no_manual',
        'no_do',
        'tgl_jt',
    ],
    
    'LK-000065' : [
        'barang_code',
        'barang_name',
        'kima',
        'satuan_1',
        'rasio_2',
        'karton_mt',

        'faktur_date',
        'faktur_code',
        'faktur_type',

        'qty',
        'satuan_barang',
        'rasio_barang',
        'karton',

        'subtotal',
        'bulan',
        'tipe',

        'detail_barang_code',
        'detail_barang_name',

        'customer_code',
        'customer_name',
        'customer_address',

        'salesperson_name',
    ],
    
    'LK-000153' : [
        'distributor',
        'branch',
        'transaction_type',
        'division',
        'principal',
        'product_group_1',
        'product_group_2',
        'product_group_3',
        'brand',
        'product_code',
        'variant_code',
        'principal_product_code',
        'product_name',
        'packaging',
        'product_class',
        'customer_code',
        'customer_name',
        'customer_address',
        'area',
        'subarea',
        'channel',
        'subchannel',
        'customer_group',
        'key_account',
        'salesperson_code',
        'salesperson_name',
        'salesco_code',
        'salesco_name',
        'supervisor_code',
        'supervisor_name',
        'year_month',
        'month',
        'transaction_date',
        'week_no',
        'invoice_no',
        'sales_method',
        'selling_type',
        'qty_sold',
        'karton_utuh',
        'qty_sold_pcs',
        'freegood_pcs',
        'tonnage',
        'volume_ltr',
        'gross_amount',
        'line_discount_1',
        'line_discount_2',
        'line_discount_3',
        'line_discount_4',
        'line_discount_5',
        'total_line_discount',
        'invoice_discount_1',
        'invoice_discount_2',
        'invoice_discount_3',
        'total_invoice_discount',
        'dpp',
        'ppn',
        'ppnbm',
        'tax_3',
        'net_amount',
        'warehouse',
        'customer_po',
        'customer_join_date',
        'tax_invoice_no',
        'tax_invoice_date',
        'proforma_invoice_no',
        'proforma_invoice_date',
        'cogs',
        'case_weight_kg',
        'tsl_qty_sold_nfg',
        'tsl_conv_pcs_to_ctn',
        'tsl_tonnage_sold_fg',
        'stock_branch',
        'stock_division',
        'stock_product_group_level_3',
        'stock_product_code',
        'stock_product_name',
        'stock_packaging',
        'unit_of_measure',
        'uom_conversion',
        'stock_qty',
        'stock_qty_pcs',
        'value_selling',
        'value_buying',
        'value_average_cost',
    ],
    
    'LK-000010': [
        'agent_name',

        'customer_code',
        'customer_name',
        'customer_address',
        'customer_phone',

        'invoice_no',
        'invoice_date',

        'customer_type',
        'city',

        'item_code',
        'item_name',

        'qty',

        'discount1_pct',
        'discount2_pct',
        'discount3_pct',
        'discount4_pct',
        'discount5_pct',

        'discount6',

        'bonus_qty',

        'invoice_value',

        'salesperson_name',
    ],
    
}

ALL_FALLBACK_KEYWORDS = [item for sublist in EXPECTED_COLUMNS.values() for item in sublist]

# ======================================================
# UNMERGE EXCEL (INTELLIGENT DIRECTIONAL FILL)
# ======================================================
def unmerge_excel(input_file):
    # wb = load_workbook(input_file)
    wb = load_workbook(input_file, data_only=True)

    for ws in wb.worksheets:
        merged_ranges = list(ws.merged_cells.ranges)

        for merged_range in merged_ranges:
            min_col = merged_range.min_col
            min_row = merged_range.min_row
            
            # Ambil nilai asli sebelum di-unmerge
            value = ws.cell(row=min_row, column=min_col).value

            start_row, end_row = merged_range.min_row, merged_range.max_row
            start_col, end_col = merged_range.min_col, merged_range.max_col

            # Unmerge secara aman
            try:
                ws.unmerge_cells(str(merged_range))
            except:
                continue

            # ATURAN FILL VALUE:
            # Jika merge cell melebar ke kanan (horizontal seperti kolom PPN), sebar nilainya.
            # Jika merge cell murni ke bawah (vertikal seperti Nama Pemasok), jangan sebar ke bawahnya!
            is_horizontal_merge = end_col > start_col
            is_vertical_only = (end_row > start_row) and (start_col == end_col)

            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    if is_vertical_only and row > start_row:
                        # Biarkan sel anak vertikal kosong agar tidak melahirkan baris data duplikat
                        ws.cell(row=row, column=col).value = None
                    else:
                        ws.cell(row=row, column=col).value = value

    temp_file = input_file.replace('.xlsx', '_unmerged.xlsx')
    wb.save(temp_file)
    return temp_file

# ======================================================
# FIND HEADER ROW
# ======================================================
def find_header_row(df_preview, expected_columns):
    if not expected_columns:
        expected_columns = ALL_FALLBACK_KEYWORDS

    expected_columns = [str(x).lower().strip() for x in expected_columns]
    best_match_row = 0
    best_match_count = 0
    found_any_match = False

    for i in range(len(df_preview)):
        row_values = df_preview.iloc[i].astype(str).str.lower().str.strip().tolist()
        matched = sum(1 for expected in expected_columns if expected in row_values)

        if matched > best_match_count:
            best_match_count = matched
            best_match_row = i
            found_any_match = True

    if found_any_match:
        return best_match_row

    for i in range(len(df_preview)):
        row_list = df_preview.iloc[i].dropna().tolist()
        if len(row_list) > 2 and len(set(row_list)) == len(row_list):
            return i

    return 0

# ======================================================
# NORMALIZE SHEET
# ======================================================
# ======================================================
# NORMALIZE SHEET (FIXED END-COLUMNS DROP)
# ======================================================
# def normalize_sheet(file_path, sheet_name, report_type):

#     preview = pd.read_excel(
#         file_path,
#         sheet_name=sheet_name,
#         header=None,
#         nrows=20
#     )

#     expected = EXPECTED_COLUMNS.get(report_type, [])
#     header_row = find_header_row(preview, expected)

#     # DEBUG HEADER DETECTION
#     # print('\n================ PREVIEW =================', flush=True)

#     # # biar tabelnya kebaca jelas
#     # print(preview.head(10).to_string(), flush=True)

#     # print('\n================ HEADER =================', flush=True)
#     # print('HEADER ROW DETECTED:', header_row, flush=True)

#     df = pd.read_excel(
#         file_path,
#         sheet_name=sheet_name,
#         header=header_row
#     )

#     # DEBUG FINAL COLUMNS
#     print('\n================ COLUMNS =================', flush=True)

#     # list kolom
#     print(df.columns.tolist(), flush=True)

#     # detail per kolom biar keliatan hidden space / unnamed
#     for idx, col in enumerate(df.columns):
#         print(idx, '|', repr(col), flush=True)

#     # cleaning
#     df.columns = [
#         " ".join(
#             str(col)
#             .replace('\n', ' ')
#             .replace('\r', ' ')
#             .split()
#         ).strip()
#         for col in df.columns
#     ]

#     return df

def normalize_sheet(file_path, sheet_name, report_type, sheet_idx):
    preview = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=20)
    expected = EXPECTED_COLUMNS.get(report_type, [])
    header_row = find_header_row(preview, expected)
    
    # df = pd.read_excel(
    #     file_path,
    #     sheet_name=sheet_name,
    #     header=header_row
    # )
    
    if report_type == 'LK-000035' and sheet_idx == 0 :

        df_raw = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=None
        )

        print("HEADER ROW:", header_row)

        header1 = df_raw.iloc[header_row].fillna('')
        header2 = df_raw.iloc[header_row + 1].fillna('')

        final_columns = []

        for parent, child in zip(header1, header2):

            parent = str(parent).strip()
            child = str(child).strip()

            if parent and child:
                col_name = f"{parent} {child}"
            elif child:
                col_name = child
            else:
                col_name = parent

            final_columns.append(col_name)

        df = df_raw.iloc[header_row + 2:].copy()

        if len(final_columns) > df.shape[1]:
            final_columns = final_columns[:df.shape[1]]

        df.columns = final_columns
        
        if report_type == 'LK-000035':

            key_cols = [
                'Nama Pemasok Utama Barang',
                'Nama Penjual',
                'Tgl Faktur',
                'No. Faktur',

                'No. Pelanggan',
                'Nama Pelanggan',
                'Alamat 1 Pelanggan',
                'Kota Pelanggan',
                'Nama Tipe Pelanggan Pelanggan'
            ]

            for col in key_cols:
                if col in df.columns:
                    df[col] = df[col].ffill()
        
    else:

        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=header_row
        )

    print(f'[{sheet_name}] HEADER ROW TRACED:', header_row)

    # # =========================================================================
    # # KHUSUS LK-000035: LOGIC COUPLING / STOP FORWARD-FILL TEPAT WAKTU
    # # =========================================================================
    # if report_type == 'LK-000035':
    #     df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    #     # 1. Cari baris jangkar utama sub-header ('kuantitas', 'jumlah', dll)
    #     target_header_row = header_row
    #     for i in range(min(20, len(df_raw))):
    #         row_str = df_raw.iloc[i].astype(str).str.lower().tolist()
    #         if 'kuantitas' in row_str or 'harga satuan barang' in row_str or 'tgl faktur' in row_str:
    #             target_header_row = i
    #             break

    #     print(f'[{sheet_name}] REAL HEADER ROW FOR LK-000035:', target_header_row)

    #     start_header = max(0, target_header_row - 1) if target_header_row > 0 else 0
        
    #     # 2. Ambil area baris header bertingkat (biasanya baris bulan dan baris sub-header)
    #     df_header_area = df_raw.iloc[start_header : target_header_row + 1].copy().reset_index(drop=True)
        
    #     # 3. Ambil baris atas (baris nama Bulan seperti 'Mar') secara mandiri
    #     row_atas = df_header_area.iloc[0].tolist()
        
    #     # 4. Lakukan batasan ffill manual agar kata 'Mar' TIDAK meluber melewati batas kolomnya sendiri
    #     last_valid_month = None
    #     for idx in range(len(row_atas)):
    #         val_str = str(row_atas[idx]).strip()
    #         # Jika sel kosong/nan, isi dengan bulan terakhir yang valid (efek ffill)
    #         if val_str == 'nan' or val_str == 'None' or val_str == '':
    #             if last_valid_month is not None:
    #                 row_atas[idx] = last_valid_month
    #         else:
    #             # Jika menabrak kolom ber-header mandiri seperti 'Inc Ppn 11%', stop ffill bulan!
    #             if 'inc ppn' in val_str.lower() or 'in ctn' in val_str.lower() or 'bulan' in val_str.lower() or 'produk' in val_str.lower():
    #                 last_valid_month = None  # Reset jangkauan ffill agar tidak meluber ke kanan
    #             else:
    #                 last_valid_month = row_atas[idx] # Kunci nama bulan baru ('Mar', 'Apr', dst.)

    #     # Masukkan kembali baris atas yang sudah dijinakkan ffill-nya
    #     df_header_area.iloc[0] = row_atas

    #     # 5. Bangun nama kolom final secara berpasangan presisi
    #     final_columns = []
    #     num_cols = df_raw.shape[1]
    #     for col_idx in range(num_cols):
    #         # Ambil komponen teks di kolom tersebut dari atas ke bawah
    #         col_parts = df_header_area[col_idx].astype(str).replace('nan', '').replace('None', '').str.strip().tolist()
            
    #         cleaned_parts = []
    #         for part in col_parts:
    #             if part and (not cleaned_parts or cleaned_parts[-1] != part):
    #                 cleaned_parts.append(part)
            
    #         col_name = " ".join(cleaned_parts).strip()
    #         final_columns.append(col_name if col_name else f"Unnamed_{col_idx}")

    #     # 6. Potong data transaksi tepat di bawah target header row
    #     df = df_raw.iloc[target_header_row + 1:].copy()
        
    #     # Sinkronisasi jumlah kolom agar tidak bergeser indeks matriknya
    #     if df.shape[1] > len(final_columns):
    #         df = df.iloc[:, :len(final_columns)]
    #     elif df.shape[1] < len(final_columns):
    #         final_columns = final_columns[:df.shape[1]]
            
    #     df.columns = final_columns

    # # =========================================================================
    # # UNTUK REPORT LAIN: JALAN NORMAL TANPA SENTUH LOGIC DI ATAS SAMA SEKALI
    # # =========================================================================
    # else:
    #     df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)

    # =========================================================================
    # PROCESS CLEANING GLOBAL (UMUM)
    # =========================================================================
    # Bersihkan spasi/newline pada nama kolom
    df.columns = [" ".join(str(col).replace('\n', ' ').replace('\r', ' ').split()).strip() for col in df.columns]

    # Buang kolom kosong Unnamed asli
    valid_cols = [col for col in df.columns if col != '' and not str(col).startswith('Unnamed:')]
    df = df.loc[:, valid_cols]

    # Hanya drop baris yang kosong total (Menyelamatkan kolom ujung kanan)
    df = df.dropna(how='all')
    df = df.reset_index(drop=True)

    # Skenario pembulatan khusus LK-000042
    if report_type == 'LK-000042':
        columns_lower = [str(col).lower().strip() for col in df.columns]
        if 'qty in ct' in columns_lower:
            original_col = df.columns[columns_lower.index('qty in ct')]
            df[original_col] = pd.to_numeric(df[original_col], errors='coerce').round().astype('Int64')

    return df

# ======================================================
# MAIN RUNNER
# ======================================================
def run():
    if len(sys.argv) < 3:
        print("Usage: python script.py <input_file.xlsx> <output_file.xlsx> [optional_report_type]")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]

    unmerged_file = unmerge_excel(input_file)
    excel = pd.ExcelFile(unmerged_file)
    
    if len(sys.argv) >= 4:

        passed_type = str(sys.argv[3]).upper().strip()

        print("PASSED TYPE:", passed_type, flush=True)

        # NORMALIZE FORMAT
        only_number = passed_type.replace('LK-', '').strip()

        report_type = f'LK-{only_number.zfill(6)}'

    else:
        report_type = detect_report(os.path.basename(input_file))

    print('FINAL DETECTED REPORT TYPE:', report_type)
    processed = {}

    for idx, sheet in enumerate(excel.sheet_names):
        # 'LK-000035', 'LK-000042', masukin ke skip lagi nanti
        if report_type in ['LK-000105', 'LK-000143', 'LK-000021', 'LK-000075', 'LK-000124'] and idx == 0:
            print(f'Skip first sheet: {sheet}')
            continue

        try:
            df = normalize_sheet(unmerged_file, sheet, report_type, idx)
            if df.empty or len(df.columns) == 0:
                print(f'Skip empty/broken sheet: {sheet}')
                continue

            processed[sheet] = df
            print(f'Success sheet: {sheet}')
        except Exception as e:
            print(f'Error sheet {sheet}: {e}')

    if os.path.exists(unmerged_file):
        os.remove(unmerged_file)

    if len(processed) == 0:
        raise Exception('Tidak ada sheet berhasil diproses')

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, sheet_df in processed.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            for column_cells in worksheet.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                worksheet.column_dimensions[column].width = max_length + 3

    print(f'NORMALIZE SUCCESS: {output_file}')

if __name__ == '__main__':
    run()
